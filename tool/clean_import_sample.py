"""Clean the source Excel export and emit an importer-ready workbook.

The source file is the 1Money export the user shared. The app's importer
(`lib/features/import/...`) expects sheets named Expenses / Income /
Transfers with the same column layout produced by 1Money. This script
rewrites the data so:

* Category names are mapped to the app's canonical defaults whenever
  possible (CategoryResolver matches case-insensitive on trimmed name).
* Account names are normalised to consistent Title Case.
* Common typos are fixed in tags + comments.
* Whitespace is trimmed everywhere.
* Rows are sorted newest-first for a more human-friendly preview.

Run:
    python3 tool/clean_import_sample.py \
        --src "/Users/rohith/Downloads/2026_05_09_01_43_24_529254 (1).xlsx" \
        --out "/Users/rohith/Downloads/money_manager_import_sample_cleaned.xlsx"
"""

from __future__ import annotations

import argparse
import re
import shutil
import tempfile
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ───────────────────────── canonical mappings ─────────────────────────────────

EXPENSE_CATEGORY_MAP: dict[str, str] = {
    "food": "Food & Dining",
    "groceries": "Food & Dining",
    "entertainment": "Entertainment",
    "education": "Education",
    "health": "Health",
    "trips": "Travel",
    "rent": "Rent & Housing",
    "home": "Rent & Housing",
    "recharge": "Utilities",
    "phone emi": "Subscriptions",
    "insurance": "Subscriptions",
    "credit card": "Other",
    "investment": "Other",
    "lended or loan": "Other",
    "lend money": "Other",
    "personal": "Other",
    "family": "Other",
    "relatives": "Other",
    "friends": "Other",
    "encah": "Other",
    "encash": "Other",
    "other": "Other",
}

INCOME_CATEGORY_MAP: dict[str, str] = {
    "salary": "Salary",
    "lend": "Other",
    "credit card": "Other",
    "cashback": "Other",
    "encash": "Other",
    "i owes": "Other",
    "gift": "Other",
    "settle up": "Other",
    "other": "Other",
}

ACCOUNT_MAP: dict[str, str] = {
    "main": "Main",
    "hdfc credit card": "HDFC Credit Card",
    "utkarsh creditcard": "Utkarsh Credit Card",
    "utkarsh credit card": "Utkarsh Credit Card",
    "cash on hand": "Cash on Hand",
    "cred wallet": "Cred Wallet",
}

# Tag and comment text fixes. Keys are case-insensitive whole words.
TYPO_FIXES: dict[str, str] = {
    "encah": "Encash",
    "wheet": "Wheat",
    "biriyani": "Biryani",
    "dosha": "Dosa",
    "porota": "Porotta",
    "vadapav": "Vada Pav",
    "icecream": "Ice Cream",
    "haircut": "Haircut",
    "hair cut": "Haircut",
    "pani poori": "Pani Puri",
    "pani puri": "Pani Puri",
    "lemon juice": "Lemon Juice",
    "youtube premium": "YouTube Premium",
    "term insurance": "Term Insurance",
}

# Tag tokens that should remain exactly lower or upper case after Title casing.
SPECIAL_TAG_CASE: dict[str, str] = {
    "Kfc": "KFC",
    "Youtube": "YouTube",
    "Hdfc": "HDFC",
}

# ──────────────────────────── helpers ─────────────────────────────────────────


def _norm_key(value: str | None) -> str:
    return (value or "").strip().lower()


def map_account(name: str) -> str:
    return ACCOUNT_MAP.get(_norm_key(name), (name or "").strip())


def map_expense_category(name: str) -> str:
    key = _norm_key(name)
    return EXPENSE_CATEGORY_MAP.get(key, (name or "").strip() or "Other")


def map_income_category(name: str) -> str:
    key = _norm_key(name)
    return INCOME_CATEGORY_MAP.get(key, (name or "").strip() or "Other")


def fix_typos(text: str | None) -> str:
    if not text:
        return ""
    cleaned = re.sub(r"\s+", " ", str(text)).strip()
    for bad, good in TYPO_FIXES.items():
        cleaned = re.sub(rf"\b{re.escape(bad)}\b", good, cleaned, flags=re.IGNORECASE)
    return cleaned


def normalise_tag(tag: str) -> str:
    cleaned = re.sub(r"\s+", " ", tag).strip()
    if not cleaned:
        return ""
    cleaned = fix_typos(cleaned).title()
    for bad, good in SPECIAL_TAG_CASE.items():
        cleaned = re.sub(rf"\b{re.escape(bad)}\b", good, cleaned)
    return cleaned


def normalise_tags(raw: str | None) -> str:
    if not raw:
        return ""
    parts = [normalise_tag(t) for t in re.split(r"[,;|]", str(raw))]
    deduped: list[str] = []
    seen: set[str] = set()
    for part in parts:
        if part and part not in seen:
            deduped.append(part)
            seen.add(part)
    return ", ".join(deduped)


def coerce_date(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def coerce_amount(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


# ──────────────────────────── core cleaning ───────────────────────────────────


EXPENSE_HEADERS = [
    "Date and time",
    "Category",
    "Account",
    "Amount in default currency",
    "Default currency",
    "Amount in account currency",
    "Account currency",
    "Transaction amount in transaction currency",
    "Transaction currency",
    "Tags",
    "Comment",
]

INCOME_HEADERS = list(EXPENSE_HEADERS)

TRANSFER_HEADERS = [
    "Date and time",
    "Outgoing",
    "Incoming",
    "Amount in outgoing currency",
    "Outgoing currency",
    "Amount in incoming currency",
    "Incoming currency",
    "Comment",
]


def clean_transactions(rows, *, is_income: bool) -> tuple[list, Counter]:
    cleaned = []
    cat_counter: Counter = Counter()
    for r in rows:
        if not any(c not in (None, "") for c in r):
            continue
        date = coerce_date(r[0])
        if date is None:
            continue
        category_raw = (r[1] or "").strip()
        account_raw = (r[2] or "").strip()
        amount_default = coerce_amount(r[3])
        if amount_default is None or amount_default == 0:
            continue
        currency_default = (r[4] or "INR").strip().upper() or "INR"
        amount_account = coerce_amount(r[5]) or amount_default
        currency_account = (r[6] or currency_default).strip().upper() or currency_default
        amount_txn = coerce_amount(r[7])
        currency_txn = (r[8] or "").strip().upper()
        tags = normalise_tags(r[9])
        comment = fix_typos(r[10])

        category = (
            map_income_category(category_raw)
            if is_income
            else map_expense_category(category_raw)
        )
        account = map_account(account_raw)
        cat_counter[category] += 1

        cleaned.append(
            [
                date,
                category,
                account,
                round(amount_default, 2),
                currency_default,
                round(amount_account, 2),
                currency_account,
                round(amount_txn, 2) if amount_txn is not None else "",
                currency_txn,
                tags,
                comment,
            ]
        )

    cleaned.sort(key=lambda row: row[0], reverse=True)
    return cleaned, cat_counter


def clean_transfers(rows) -> list:
    cleaned = []
    for r in rows:
        if not any(c not in (None, "") for c in r):
            continue
        date = coerce_date(r[0])
        if date is None:
            continue
        outgoing = map_account((r[1] or "").strip())
        incoming = map_account((r[2] or "").strip())
        if not outgoing or not incoming:
            continue
        amount_out = coerce_amount(r[3])
        if amount_out is None or amount_out == 0:
            continue
        currency_out = (r[4] or "INR").strip().upper() or "INR"
        amount_in = coerce_amount(r[5]) or amount_out
        currency_in = (r[6] or currency_out).strip().upper() or currency_out
        comment = fix_typos(r[7])

        cleaned.append(
            [
                date,
                outgoing,
                incoming,
                round(amount_out, 2),
                currency_out,
                round(amount_in, 2),
                currency_in,
                comment,
            ]
        )

    cleaned.sort(key=lambda row: row[0], reverse=True)
    return cleaned


# ──────────────────────────── workbook writing ────────────────────────────────


def _write_sheet(ws, headers, rows, *, banner: str) -> None:
    ws.append([banner] + [None] * (len(headers) - 1))
    ws.append(headers)
    for r in rows:
        ws.append(r)

    bold_blue = Font(bold=True, color="FFFFFFFF")
    fill = PatternFill("solid", fgColor="FF0052FF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = bold_blue
        cell.fill = fill

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            18, len(header) + 2
        )

    for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
        date_cell = row[0]
        if isinstance(date_cell.value, datetime):
            date_cell.number_format = "yyyy-mm-dd"


def build_workbook(src: Path) -> tuple[openpyxl.Workbook, dict]:
    src_wb = openpyxl.load_workbook(src, data_only=True)

    src_expenses = list(src_wb["Expenses"].iter_rows(values_only=True))[2:]
    src_income = list(src_wb["Income"].iter_rows(values_only=True))[2:]
    src_transfers = list(src_wb["Transfers"].iter_rows(values_only=True))[2:]

    expenses, expense_cats = clean_transactions(src_expenses, is_income=False)
    income, income_cats = clean_transactions(src_income, is_income=True)
    transfers = clean_transfers(src_transfers)

    out = openpyxl.Workbook()
    out.remove(out.active)
    _write_sheet(
        out.create_sheet("Expenses"),
        EXPENSE_HEADERS,
        expenses,
        banner="expenses list",
    )
    _write_sheet(
        out.create_sheet("Income"),
        INCOME_HEADERS,
        income,
        banner="income list",
    )
    _write_sheet(
        out.create_sheet("Transfers"),
        TRANSFER_HEADERS,
        transfers,
        banner="transfers list",
    )

    summary = {
        "expenses": len(expenses),
        "income": len(income),
        "transfers": len(transfers),
        "expense_categories": dict(expense_cats),
        "income_categories": dict(income_cats),
        "earliest": min(
            (r[0] for r in expenses + income + transfers if isinstance(r[0], datetime)),
            default=None,
        ),
        "latest": max(
            (r[0] for r in expenses + income + transfers if isinstance(r[0], datetime)),
            default=None,
        ),
    }
    return out, summary


_ABSOLUTE_REL_TARGET = re.compile(r'Target="/xl/')

# Self-closing inline-string cells (`<c r="A1" t="inlineStr" />`) crash the
# Dart `excel` package because its `case 'inlineStr':` branch unconditionally
# does `findAllElements('t').first`. Excel treats absent cells as empty
# anyway, so we drop those empty placeholders entirely.
_EMPTY_INLINE_STR_CELL = re.compile(
    r'<c\b[^>]*\bt="inlineStr"[^>]*/>'
)


def _normalise_for_dart_parser(path: Path) -> None:
    """Patch the workbook so the Dart ``excel`` package can read it.

    Two fixes are applied:

    * Worksheet relationship targets are rewritten from
      ``/xl/worksheets/sheetN.xml`` to ``worksheets/sheetN.xml`` because the
      Dart parser looks them up via ``xl/$target``.
    * Empty self-closing ``<c .../>`` inline-string placeholders are stripped
      from each worksheet so the parser doesn't crash on missing ``<t>``.
    """
    tmp = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    try:
        with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(
            tmp, "w", zipfile.ZIP_DEFLATED
        ) as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename == "xl/_rels/workbook.xml.rels":
                    text = data.decode("utf-8")
                    text = _ABSOLUTE_REL_TARGET.sub('Target="', text)
                    data = text.encode("utf-8")
                elif item.filename.startswith(
                    "xl/worksheets/"
                ) and item.filename.endswith(".xml"):
                    text = data.decode("utf-8")
                    text = _EMPTY_INLINE_STR_CELL.sub("", text)
                    data = text.encode("utf-8")
                dst.writestr(item, data)
        shutil.move(tmp, path)
    finally:
        if tmp.exists():
            tmp.unlink()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()

    wb, summary = build_workbook(args.src)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    _normalise_for_dart_parser(args.out)

    print(f"Wrote: {args.out}")
    print(f"  Expenses: {summary['expenses']}")
    print(f"  Income:   {summary['income']}")
    print(f"  Transfers:{summary['transfers']}")
    print(f"  Range:    {summary['earliest']} → {summary['latest']}")
    print(f"  Expense category distribution: {summary['expense_categories']}")
    print(f"  Income category distribution:  {summary['income_categories']}")


if __name__ == "__main__":
    main()
